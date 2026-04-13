#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
爬虫需求总结工具 - 创建包含10个爬虫需求的Excel文件
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
import os

def create_crawler_requirements():
    """创建爬虫需求数据"""
    requirements = [
        {
            "需求编号": "CR001",
            "需求名称": "电商价格监控爬虫",
            "应用场景": "价格竞争分析、市场监控",
            "目标网站/平台": "淘宝、京东、拼多多、亚马逊",
            "主要数据类型": "价格、销量、评价、库存",
            "技术复杂度": "中",
            "反爬虫难度": "高",
            "数据更新频率": "每天多次",
            "预计数据量": "大",
            "数据用途": "价格监控、竞争分析、动态定价",
            "关键技术要点": "动态页面渲染、登录验证、IP代理池、验证码识别",
            "建议工具/框架": "Selenium + Scrapy + Redis",
            "法律合规风险": "中",
            "备注": "需要处理JavaScript渲染和频繁的网站改版"
        },
        {
            "需求编号": "CR002",
            "需求名称": "新闻资讯聚合爬虫",
            "应用场景": "新闻聚合、舆情监控",
            "目标网站/平台": "新浪新闻、腾讯新闻、网易新闻、头条",
            "主要数据类型": "标题、正文、时间、来源、作者",
            "技术复杂度": "低",
            "反爬虫难度": "低",
            "数据更新频率": "实时/每15分钟",
            "预计数据量": "大",
            "数据用途": "新闻聚合、舆情分析、热点追踪",
            "关键技术要点": "HTML解析、去重策略、增量更新",
            "建议工具/框架": "Scrapy + BeautifulSoup",
            "法律合规风险": "低",
            "备注": "需注意版权问题，可添加RSS源支持"
        },
        {
            "需求编号": "CR003",
            "需求名称": "社交媒体情感分析爬虫",
            "应用场景": "品牌监控、市场调研",
            "目标网站/平台": "微博、知乎、小红书、抖音",
            "主要数据类型": "评论、点赞数、转发数、用户信息",
            "技术复杂度": "高",
            "反爬虫难度": "高",
            "数据更新频率": "实时",
            "预计数据量": "极大",
            "数据用途": "情感分析、品牌声誉监控、用户洞察",
            "关键技术要点": "API调用限制、登录状态维护、反爬虫对抗",
            "建议工具/框架": "Requests + Selenium + 机器学习库",
            "法律合规风险": "高",
            "备注": "需严格遵守隐私政策，建议数据匿名化处理"
        },
        {
            "需求编号": "CR004",
            "需求名称": "招聘网站职位信息爬虫",
            "应用场景": "人才市场分析、职位监控",
            "目标网站/平台": "前程无忧、智联招聘、BOSS直聘",
            "主要数据类型": "职位名称、薪资范围、要求、公司信息",
            "技术复杂度": "中",
            "反爬虫难度": "中",
            "数据更新频率": "每天",
            "预计数据量": "中",
            "数据用途": "人才市场分析、薪资调研、技能需求分析",
            "关键技术要点": "分页处理、搜索条件模拟、数据去重",
            "建议工具/框架": "Scrapy + MongoDB",
            "法律合规风险": "中",
            "备注": "需注意避免对招聘网站服务器造成过大压力"
        },
        {
            "需求编号": "CR005",
            "需求名称": "房地产市场价格爬虫",
            "应用场景": "房价监控、市场分析",
            "目标网站/平台": "链家、贝壳、安居客、房天下",
            "主要数据类型": "价格、面积、户型、位置、小区信息",
            "技术复杂度": "中",
            "反爬虫难度": "中",
            "数据更新频率": "每周",
            "预计数据量": "中",
            "数据用途": "房价趋势分析、投资参考、区域比较",
            "关键技术要点": "地图坐标解析、价格计算、数据校验",
            "建议工具/框架": "Scrapy + PostgreSQL + GeoDjango",
            "法律合规风险": "低",
            "备注": "可结合地理位置信息进行房价热力图分析"
        },
        {
            "需求编号": "CR006",
            "需求名称": "股票金融数据爬虫",
            "应用场景": "投资分析、市场研究",
            "目标网站/平台": "东方财富、新浪财经、雪球、Yahoo Finance",
            "主要数据类型": "股价、成交量、财务数据、新闻",
            "技术复杂度": "低",
            "反爬虫难度": "低",
            "数据更新频率": "实时/每5分钟",
            "预计数据量": "大",
            "数据用途": "投资分析、量化交易、市场监控",
            "关键技术要点": "API调用、数据清洗、实时更新",
            "建议工具/框架": "Requests + Pandas",
            "法律合规风险": "低",
            "备注": "注意数据延迟问题，推荐使用官方API优先"
        },
        {
            "需求编号": "CR007",
            "需求名称": "学术论文文献爬虫",
            "应用场景": "学术研究、文献分析",
            "目标网站/平台": "知网、万方、Google Scholar、arXiv",
            "主要数据类型": "标题、作者、摘要、关键词、引用",
            "技术复杂度": "低",
            "反爬虫难度": "低",
            "数据更新频率": "每月",
            "预计数据量": "大",
            "数据用途": "文献综述、研究趋势分析、作者网络分析",
            "关键技术要点": "PDF解析、引用关系提取、数据标准化",
            "建议工具/框架": "Scrapy + PDF解析库",
            "法律合规风险": "中",
            "备注": "需注意版权和合理使用原则"
        },
        {
            "需求编号": "CR008",
            "需求名称": "旅游网站酒店价格爬虫",
            "应用场景": "旅游规划、价格比较",
            "目标网站/平台": "携程、去哪儿、Booking.com、Airbnb",
            "主要数据类型": "价格、房型、评分、位置、设施",
            "技术复杂度": "高",
            "反爬虫难度": "高",
            "数据更新频率": "每天",
            "预计数据量": "大",
            "数据用途": "价格比较、最优预订时间分析",
            "关键技术要点": "动态定价策略解析、用户模拟、多条件查询",
            "建议工具/框架": "Selenium + Scrapy + 分布式爬虫",
            "法律合规风险": "中",
            "备注": "需注意预订政策变化和价格波动规律"
        },
        {
            "需求编号": "CR009",
            "需求名称": "视频平台热门内容爬虫",
            "应用场景": "内容趋势分析、流行文化研究",
            "目标网站/平台": "B站、抖音、YouTube、腾讯视频",
            "主要数据类型": "视频信息、播放量、评论、弹幕",
            "技术复杂度": "高",
            "反爬虫难度": "高",
            "数据更新频率": "每天",
            "预计数据量": "极大",
            "数据用途": "内容趋势分析、用户偏好研究",
            "关键技术要点": "视频流处理、大规模数据存储、实时分析",
            "建议工具/框架": "分布式爬虫 + Kafka + Spark",
            "法律合规风险": "高",
            "备注": "需处理大量多媒体数据，存储成本较高"
        },
        {
            "需求编号": "CR010",
            "需求名称": "政府公开数据爬虫",
            "应用场景": "公共数据收集、政策研究",
            "目标网站/平台": "各地政府数据开放平台、统计局网站",
            "主要数据类型": "统计数据、报告、政策文件、表格",
            "技术复杂度": "低",
            "反爬虫难度": "低",
            "数据更新频率": "每月/每季度",
            "预计数据量": "中",
            "数据用途": "政策分析、经济研究、社会调查",
            "关键技术要点": "文档格式转换、数据清洗、标准化",
            "建议工具/框架": "Requests + Pandas + 文档解析库",
            "法律合规风险": "低",
            "备注": "公开数据，合规风险低，适合初学者练习"
        }
    ]
    return requirements

def apply_styling(worksheet):
    """应用样式到工作表"""
    
    # 定义样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    normal_font = Font(size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 应用表头样式
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # 应用数据行样式
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.font = normal_font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
    
    # 设置列宽
    column_widths = {
        'A': 10,  # 需求编号
        'B': 20,  # 需求名称
        'C': 15,  # 应用场景
        'D': 25,  # 目标网站/平台
        'E': 15,  # 主要数据类型
        'F': 12,  # 技术复杂度
        'G': 12,  # 反爬虫难度
        'H': 12,  # 数据更新频率
        'I': 10,  # 预计数据量
        'J': 20,  # 数据用途
        'K': 25,  # 关键技术要点
        'L': 15,  # 建议工具/框架
        'M': 12,  # 法律合规风险
        'N': 30,  # 备注
    }
    
    for col, width in column_widths.items():
        worksheet.column_dimensions[col].width = width
    
    # 添加条件格式：技术复杂度着色
    for col in ['F', 'G', 'M']:  # 技术复杂度、反爬虫难度、法律合规风险
        col_letter = get_column_letter(list(column_widths.keys()).index(col) + 1)
        
        # 高难度 - 红色
        worksheet.conditional_formatting.add(
            f'{col_letter}2:{col_letter}11',
            CellIsRule(operator='equal', formula=['"高"'], fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
        )
        
        # 中难度 - 黄色
        worksheet.conditional_formatting.add(
            f'{col_letter}2:{col_letter}11',
            CellIsRule(operator='equal', formula=['"中"'], fill=PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"))
        )
        
        # 低难度 - 绿色
        worksheet.conditional_formatting.add(
            f'{col_letter}2:{col_letter}11',
            CellIsRule(operator='equal', formula=['"低"'], fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
        )
    
    # 冻结窗格（冻结表头行）
    worksheet.freeze_panes = 'A2'

def create_summary_sheet(workbook, df):
    """创建汇总统计工作表"""
    summary_sheet = workbook.create_sheet(title='汇总统计')
    
    # 设置标题
    summary_sheet.title = '汇总统计'
    
    # 写入主标题
    summary_sheet['A1'] = '爬虫需求分析汇总报告'
    summary_sheet['A1'].font = Font(bold=True, size=16, color="366092")
    summary_sheet.merge_cells('A1:D1')
    
    # 写入统计日期
    from datetime import datetime
    summary_sheet['A3'] = f'生成时间：{datetime.now().strftime("%Y年%m月%d日 %H:%M")}'
    summary_sheet['A3'].font = Font(italic=True, size=10)
    
    # 写入总体统计
    summary_sheet['A5'] = '总体统计'
    summary_sheet['A5'].font = Font(bold=True, size=14)
    
    summary_sheet['A6'] = '总需求数量：'
    summary_sheet['B6'] = len(df)
    summary_sheet['A7'] = '平均技术复杂度：'
    summary_sheet['A8'] = '平均反爬虫难度：'
    summary_sheet['A9'] = '平均法律合规风险：'
    
    # 技术复杂度统计
    summary_sheet['A11'] = '技术复杂度分布'
    summary_sheet['A11'].font = Font(bold=True, size=14)
    
    summary_sheet['A12'] = '复杂度级别'
    summary_sheet['B12'] = '数量'
    summary_sheet['C12'] = '占比'
    
    tech_stats = df['技术复杂度'].value_counts()
    row = 13
    for level, count in tech_stats.items():
        summary_sheet[f'A{row}'] = level
        summary_sheet[f'B{row}'] = count
        summary_sheet[f'C{row}'] = f"{count/len(df)*100:.1f}%"
        row += 1
    
    # 反爬虫难度统计
    summary_sheet['E11'] = '反爬虫难度分布'
    summary_sheet['E11'].font = Font(bold=True, size=14)
    
    summary_sheet['E12'] = '难度级别'
    summary_sheet['F12'] = '数量'
    summary_sheet['G12'] = '占比'
    
    anti_stats = df['反爬虫难度'].value_counts()
    row = 13
    for level, count in anti_stats.items():
        summary_sheet[f'E{row}'] = level
        summary_sheet[f'F{row}'] = count
        summary_sheet[f'G{row}'] = f"{count/len(df)*100:.1f}%"
        row += 1
    
    # 应用样式到汇总表
    for row in summary_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.row == 1:
                cell.font = Font(bold=True, size=16)
            elif cell.row in [5, 11, 12]:
                cell.font = Font(bold=True)
    
    # 设置列宽
    summary_sheet.column_dimensions['A'].width = 20
    summary_sheet.column_dimensions['B'].width = 10
    summary_sheet.column_dimensions['C'].width = 10
    summary_sheet.column_dimensions['E'].width = 15
    summary_sheet.column_dimensions['F'].width = 10
    summary_sheet.column_dimensions['G'].width = 10

def main():
    """主函数"""
    print("开始创建爬虫需求总结Excel文件...")
    
    # 创建需求数据
    requirements = create_crawler_requirements()
    df = pd.DataFrame(requirements)
    
    # 创建Excel文件
    excel_file = "爬虫需求总结.xlsx"
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        # 写入主数据
        df.to_excel(writer, sheet_name='爬虫需求详情', index=False)
        
        # 获取工作簿和工作表
        workbook = writer.book
        worksheet = writer.sheets['爬虫需求详情']
        
        # 应用样式
        apply_styling(worksheet)
        
        # 创建汇总统计
        create_summary_sheet(workbook, df)
    
    print(f"✓ Excel文件已成功创建：{excel_file}")
    print(f"✓ 包含{len(requirements)}个详细的爬虫需求分析")
    print(f"✓ 文件大小：{os.path.getsize(excel_file)/1024:.1f} KB")
    print("\n工作表说明：")
    print("1. '爬虫需求详情' - 包含10个爬虫需求的详细信息")
    print("2. '汇总统计' - 包含各类统计信息和分布情况")
    
    # 显示数据预览
    print("\n需求概览：")
    for i, req in enumerate(requirements, 1):
        print(f"{i:2d}. {req['需求名称']} - {req['应用场景']}")

if __name__ == "__main__":
    main()