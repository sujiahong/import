#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel导出工具 - 将爬取的数据导出为Excel格式
"""

import pandas as pd
import os
from datetime import datetime
from typing import Dict, List, Any, Optional
from pathlib import Path

from config.settings import EXCEL_DIR, PROJECT_ROOT


class ExcelExporter:
    """Excel导出工具类"""
    
    def __init__(self, output_dir: Optional[Path] = None):
        """
        初始化导出器
        
        Args:
            output_dir: 输出目录，默认为配置中的EXCEL_DIR
        """
        self.output_dir = output_dir or EXCEL_DIR
        self.output_dir.mkdir(exist_ok=True)
        
        # 默认样式配置
        self.styles = {
            "header": {
                "font": {"bold": True, "color": "FFFFFF"},
                "fill": {"start_color": "366092", "end_color": "366092", "fill_type": "solid"},
                "alignment": {"horizontal": "center", "vertical": "center"}
            },
            "date": {
                "number_format": "yyyy-mm-dd hh:mm:ss"
            },
            "number": {
                "number_format": "#,##0"
            },
            "currency": {
                "number_format": "¥#,##0.00"
            }
        }
    
    def export_to_excel(self, data: List[Dict], filename: str, 
                       sheet_name: str = "数据", 
                       include_stats: bool = True,
                       include_charts: bool = False) -> Path:
        """
        导出数据到Excel文件
        
        Args:
            data: 要导出的数据列表
            filename: 输出文件名
            sheet_name: 工作表名称
            include_stats: 是否包含统计信息
            include_charts: 是否包含图表
            
        Returns:
            输出文件的路径
        """
        if not data:
            raise ValueError("没有数据需要导出")
        
        # 确保文件扩展名
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        filepath = self.output_dir / filename
        
        # 创建DataFrame
        df = pd.DataFrame(data)
        
        # 创建Excel写入器
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # 写入主数据
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # 获取工作簿和工作表
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # 应用基本样式
            self._apply_basic_styling(worksheet, df)
            
            # 添加统计信息
            if include_stats:
                self._add_statistics_sheet(workbook, df, sheet_name)
            
            # 添加图表（可选）
            if include_charts and len(df) > 0:
                self._add_charts(workbook, df, sheet_name)
        
        print(f"数据已导出到: {filepath}")
        print(f"总记录数: {len(df)}")
        
        return filepath
    
    def export_crawler_data(self, crawler_name: str, data: List[Dict], 
                           crawler_stats: Optional[Dict] = None) -> Path:
        """
        专门为爬虫数据设计的导出方法
        
        Args:
            crawler_name: 爬虫名称
            data: 爬取的数据
            crawler_stats: 爬虫统计信息
            
        Returns:
            输出文件的路径
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{crawler_name}_{timestamp}.xlsx"
        
        filepath = self.output_dir / filename
        
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # 1. 主数据工作表
            if data:
                df_data = pd.DataFrame(data)
                df_data.to_excel(writer, sheet_name="爬取数据", index=False)
                
                # 应用样式
                worksheet_data = writer.sheets["爬取数据"]
                self._apply_basic_styling(worksheet_data, df_data)
            
            # 2. 统计信息工作表
            workbook = writer.book
            self._add_crawler_statistics_sheet(workbook, data, crawler_stats, crawler_name)
            
            # 3. 数据概览工作表
            if data:
                self._add_data_overview_sheet(workbook, data, crawler_name)
        
        print(f"爬虫数据已导出到: {filepath}")
        if data:
            print(f"爬取记录数: {len(data)}")
        
        return filepath
    
    def _apply_basic_styling(self, worksheet, df: pd.DataFrame):
        """应用基本样式到工作表"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 定义样式
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        normal_font = Font(size=10)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 应用表头样式
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # 应用数据行样式
        for row in worksheet.iter_rows(min_row=2, max_row=len(df)+1):
            for cell in row:
                cell.font = normal_font
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)
        
        # 自动调整列宽
        for i, column in enumerate(df.columns, 1):
            column_letter = get_column_letter(i)
            
            # 计算最大宽度
            max_length = 0
            column_name = str(column)
            max_length = max(max_length, len(column_name))
            
            # 检查数据中的最大长度
            column_data = df[column].astype(str)
            for cell_value in column_data:
                max_length = max(max_length, len(str(cell_value)))
            
            # 设置列宽（限制最大宽度）
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 冻结窗格（冻结表头）
        worksheet.freeze_panes = 'A2'
    
    def _add_statistics_sheet(self, workbook, df: pd.DataFrame, source_sheet_name: str):
        """添加统计信息工作表"""
        from openpyxl.styles import Font, Alignment
        
        # 创建统计工作表
        stats_sheet = workbook.create_sheet(title="统计信息")
        
        # 添加标题
        stats_sheet.title = "统计信息"
        stats_sheet['A1'] = f"数据统计 - {source_sheet_name}"
        stats_sheet['A1'].font = Font(bold=True, size=16, color="366092")
        stats_sheet.merge_cells('A1:D1')
        
        # 添加生成时间
        stats_sheet['A3'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        stats_sheet['A3'].font = Font(italic=True, size=10)
        
        # 基本统计
        stats_sheet['A5'] = "基本统计信息"
        stats_sheet['A5'].font = Font(bold=True, size=14)
        
        stats_sheet['A6'] = "总记录数:"
        stats_sheet['B6'] = len(df)
        
        stats_sheet['A7'] = "列数:"
        stats_sheet['B7'] = len(df.columns)
        
        # 各列统计
        row = 10
        stats_sheet[f'A{row}'] = "列统计"
        stats_sheet[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        stats_sheet[f'A{row}'] = "列名"
        stats_sheet[f'B{row}'] = "数据类型"
        stats_sheet[f'C{row}'] = "非空值数"
        stats_sheet[f'D{row}'] = "空值数"
        stats_sheet[f'E{row}'] = "唯一值数"
        
        header_row = row
        for cell in stats_sheet[header_row]:
            cell.font = Font(bold=True)
        
        row += 1
        
        for column in df.columns:
            col_data = df[column]
            non_null_count = col_data.notna().sum()
            null_count = col_data.isna().sum()
            unique_count = col_data.nunique()
            
            stats_sheet[f'A{row}'] = str(column)
            stats_sheet[f'B{row}'] = str(col_data.dtype)
            stats_sheet[f'C{row}'] = non_null_count
            stats_sheet[f'D{row}'] = null_count
            stats_sheet[f'E{row}'] = unique_count
            
            row += 1
        
        # 设置列宽
        stats_sheet.column_dimensions['A'].width = 20
        stats_sheet.column_dimensions['B'].width = 15
        stats_sheet.column_dimensions['C'].width = 12
        stats_sheet.column_dimensions['D'].width = 12
        stats_sheet.column_dimensions['E'].width = 12
        
        # 居中对齐
        for row in stats_sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _add_crawler_statistics_sheet(self, workbook, data: List[Dict], 
                                     crawler_stats: Optional[Dict], crawler_name: str):
        """添加爬虫统计信息工作表"""
        from openpyxl.styles import Font, Alignment
        
        stats_sheet = workbook.create_sheet(title="爬虫统计")
        
        # 标题
        stats_sheet['A1'] = f"爬虫统计 - {crawler_name}"
        stats_sheet['A1'].font = Font(bold=True, size=16, color="366092")
        stats_sheet.merge_cells('A1:E1')
        
        # 生成时间
        stats_sheet['A3'] = f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        stats_sheet['A3'].font = Font(italic=True, size=10)
        
        # 数据统计
        row = 5
        stats_sheet[f'A{row}'] = "数据统计"
        stats_sheet[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        stats_sheet[f'A{row}'] = "总数据记录数:"
        stats_sheet[f'B{row}'] = len(data)
        row += 1
        
        if data:
            # 计算字段统计
            fields = set()
            for item in data:
                fields.update(item.keys())
            
            stats_sheet[f'A{row}'] = "数据字段数:"
            stats_sheet[f'B{row}'] = len(fields)
            row += 2
            
            # 字段列表
            stats_sheet[f'A{row}'] = "数据字段列表:"
            stats_sheet[f'A{row}'].font = Font(bold=True)
            row += 1
            
            for field in sorted(fields):
                stats_sheet[f'A{row}'] = f"  • {field}"
                row += 1
        
        # 爬虫性能统计
        if crawler_stats:
            row += 2
            stats_sheet[f'A{row}'] = "爬虫性能统计"
            stats_sheet[f'A{row}'].font = Font(bold=True, size=14)
            row += 2
            
            for key, value in crawler_stats.items():
                if key not in ['start_time', 'end_time']:
                    stats_sheet[f'A{row}'] = f"{key}:"
                    stats_sheet[f'B{row}'] = value
                    row += 1
            
            # 计算成功率
            if 'total_requests' in crawler_stats and 'successful_requests' in crawler_stats:
                if crawler_stats['total_requests'] > 0:
                    success_rate = (crawler_stats['successful_requests'] / crawler_stats['total_requests']) * 100
                    stats_sheet[f'A{row}'] = "请求成功率:"
                    stats_sheet[f'B{row}'] = f"{success_rate:.1f}%"
                    row += 1
        
        # 设置样式
        stats_sheet.column_dimensions['A'].width = 25
        stats_sheet.column_dimensions['B'].width = 15
        
        for row_cells in stats_sheet.iter_rows():
            for cell in row_cells:
                cell.alignment = Alignment(vertical="center")
    
    def _add_data_overview_sheet(self, workbook, data: List[Dict], crawler_name: str):
        """添加数据概览工作表"""
        from openpyxl.styles import Font, Alignment
        
        overview_sheet = workbook.create_sheet(title="数据概览")
        
        # 标题
        overview_sheet['A1'] = f"数据概览 - {crawler_name}"
        overview_sheet['A1'].font = Font(bold=True, size=16, color="366092")
        overview_sheet.merge_cells('A1:F1')
        
        # 数据样本
        row = 3
        overview_sheet[f'A{row}'] = "数据样本（前10条）"
        overview_sheet[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        # 显示前10条数据
        sample_data = data[:10]
        
        if sample_data:
            # 获取所有字段
            all_fields = set()
            for item in sample_data:
                all_fields.update(item.keys())
            
            fields = sorted(all_fields)
            
            # 写入表头
            for col_idx, field in enumerate(fields, 1):
                cell = overview_sheet.cell(row=row, column=col_idx)
                cell.value = field
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row += 1
            
            # 写入数据
            for item in sample_data:
                for col_idx, field in enumerate(fields, 1):
                    cell = overview_sheet.cell(row=row, column=col_idx)
                    value = item.get(field, "")
                    
                    # 截断过长的文本
                    if isinstance(value, str) and len(value) > 100:
                        value = value[:100] + "..."
                    
                    cell.value = value
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                
                row += 1
            
            # 设置列宽
            for col_idx, field in enumerate(fields, 1):
                max_length = len(str(field))
                for item in sample_data:
                    value = str(item.get(field, ""))
                    if len(value) > 100:
                        value = value[:100] + "..."
                    max_length = max(max_length, len(value))
                
                column_letter = chr(64 + col_idx)  # A, B, C, ...
                overview_sheet.column_dimensions[column_letter].width = min(max_length + 2, 30)
        
        # 数据摘要
        row += 2
        overview_sheet[f'A{row}'] = "数据摘要"
        overview_sheet[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        overview_sheet[f'A{row}'] = "总数据量:"
        overview_sheet[f'B{row}'] = f"{len(data)} 条记录"
        row += 1
        
        if data:
            # 计算数据大小（估算）
            import json
            data_json = json.dumps(data, ensure_ascii=False)
            data_size_mb = len(data_json.encode('utf-8')) / (1024 * 1024)
            
            overview_sheet[f'A{row}'] = "数据大小（估算）:"
            overview_sheet[f'B{row}'] = f"{data_size_mb:.2f} MB"
            row += 1
            
            # 数据时间范围（如果有时间字段）
            time_fields = ['publish_time', 'timestamp', 'crawl_time', 'time', 'date']
            for field in time_fields:
                times = [item.get(field) for item in data if item.get(field)]
                if times:
                    try:
                        # 尝试解析时间
                        valid_times = [t for t in times if t]
                        if valid_times:
                            overview_sheet[f'A{row}'] = f"时间字段 '{field}':"
                            overview_sheet[f'B{row}'] = f"{len(valid_times)} 条有效记录"
                            row += 1
                            break
                    except:
                        continue
    
    def _add_charts(self, workbook, df: pd.DataFrame, source_sheet_name: str):
        """添加图表到工作簿（基础实现）"""
        try:
            from openpyxl.chart import BarChart, Reference
            
            # 创建图表工作表
            chart_sheet = workbook.create_sheet(title="图表")
            
            # 这里可以添加具体的图表逻辑
            # 例如：创建柱状图显示数据分布
            
            # 简单示例：创建空图表工作表
            chart_sheet['A1'] = "数据图表"
            chart_sheet['A1'].font = Font(bold=True, size=14)
            
            chart_sheet['A3'] = "图表功能待实现"
            chart_sheet['A3'].font = Font(italic=True, color="666666")
            
        except ImportError:
            print("警告: openpyxl图表功能不可用，跳过图表生成")
        except Exception as e:
            print(f"图表生成失败: {e}")


def demo_excel_export():
    """演示Excel导出功能"""
    print("开始演示Excel导出功能...")
    
    # 创建示例数据
    sample_data = [
        {
            "id": 1,
            "name": "示例新闻1",
            "source": "新浪新闻",
            "category": "科技",
            "publish_time": "2024-01-15 10:30:00",
            "views": 1500,
            "content": "这是一条示例新闻内容..."
        },
        {
            "id": 2,
            "name": "示例新闻2",
            "source": "腾讯新闻",
            "category": "财经",
            "publish_time": "2024-01-15 11:45:00",
            "views": 2300,
            "content": "这是另一条示例新闻内容..."
        },
        {
            "id": 3,
            "name": "示例新闻3",
            "source": "网易新闻",
            "category": "体育",
            "publish_time": "2024-01-15 14:20:00",
            "views": 1800,
            "content": "这是第三条示例新闻内容..."
        }
    ]
    
    # 创建导出器
    exporter = ExcelExporter()
    
    try:
        # 导出数据
        filepath = exporter.export_to_excel(
            data=sample_data,
            filename="demo_export.xlsx",
            sheet_name="示例数据",
            include_stats=True,
            include_charts=False
        )
        
        print(f"示例数据已导出到: {filepath}")
        
        # 演示爬虫数据导出
        crawler_stats = {
            "total_requests": 50,
            "successful_requests": 48,
            "failed_requests": 2,
            "total_items": 3,
            "duration_seconds": 12.5
        }
        
        crawler_filepath = exporter.export_crawler_data(
            crawler_name="demo_crawler",
            data=sample_data,
            crawler_stats=crawler_stats
        )
        
        print(f"爬虫数据已导出到: {crawler_filepath}")
        
    except Exception as e:
        print(f"导出失败: {e}")


if __name__ == "__main__":
    demo_excel_export()